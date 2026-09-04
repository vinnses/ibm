#!/usr/bin/env python3
"""Extrai do arquivo INEP somente o curso de Informática Biomédica/UFPR.

O script lê uma ou mais planilhas oficiais dos Indicadores de Trajetória,
localiza a linha de cabeçalhos pelos códigos de variável e preserva todas as
colunas publicadas. A saída inclui o nome do arquivo-fonte para rastreabilidade.
"""

from __future__ import annotations

import argparse
import csv
from pathlib import Path

from openpyxl import load_workbook


CO_IES_UFPR = 571
CO_CURSO_INFORMATICA_BIOMEDICA = 1126727


def como_inteiro(valor: object) -> int | None:
    """Converte códigos que o Excel pode entregar como número ou texto."""
    if valor is None or valor == "":
        return None
    try:
        return int(float(str(valor).strip()))
    except (TypeError, ValueError):
        return None


def extrair(
    caminho: Path, ano_ingresso_min: int | None = None, ano_ingresso_max: int | None = None
) -> tuple[list[str], list[dict[str, object]]]:
    workbook = load_workbook(caminho, read_only=True, data_only=True)
    worksheet = workbook.active
    iterator = worksheet.iter_rows(values_only=True)

    cabecalho: list[str] | None = None
    for linha in iterator:
        valores = [str(valor).strip() if valor is not None else "" for valor in linha]
        if "CO_IES" in valores and "CO_CURSO" in valores:
            cabecalho = valores
            break
    if cabecalho is None:
        raise ValueError(f"Cabeçalho de variáveis não localizado em {caminho}")

    indice_ies = cabecalho.index("CO_IES")
    indice_curso = cabecalho.index("CO_CURSO")
    indice_ano_ingresso = cabecalho.index("NU_ANO_INGRESSO")
    registros: list[dict[str, object]] = []
    for linha in iterator:
        if (
            como_inteiro(linha[indice_ies]) == CO_IES_UFPR
            and como_inteiro(linha[indice_curso]) == CO_CURSO_INFORMATICA_BIOMEDICA
            and (
                ano_ingresso_min is None
                or como_inteiro(linha[indice_ano_ingresso]) >= ano_ingresso_min
            )
            and (
                ano_ingresso_max is None
                or como_inteiro(linha[indice_ano_ingresso]) <= ano_ingresso_max
            )
        ):
            registro = dict(zip(cabecalho, linha))
            registro["ARQUIVO_FONTE"] = caminho.name
            registros.append(registro)

    return cabecalho, registros


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("arquivos", nargs="+", type=Path)
    parser.add_argument("--saida", required=True, type=Path)
    parser.add_argument("--ano-ingresso-min", type=int)
    parser.add_argument("--ano-ingresso-max", type=int)
    args = parser.parse_args()

    todos: list[dict[str, object]] = []
    colunas: list[str] = []
    for arquivo in sorted(args.arquivos):
        cabecalho, registros = extrair(arquivo, args.ano_ingresso_min, args.ano_ingresso_max)
        for coluna in cabecalho:
            if coluna and coluna not in colunas:
                colunas.append(coluna)
        todos.extend(registros)
        print(f"{arquivo.name}: {len(registros)} registros")

    if "ARQUIVO_FONTE" not in colunas:
        colunas.append("ARQUIVO_FONTE")
    args.saida.parent.mkdir(parents=True, exist_ok=True)
    with args.saida.open("w", encoding="utf-8", newline="") as arquivo_saida:
        escritor = csv.DictWriter(arquivo_saida, fieldnames=colunas, extrasaction="ignore")
        escritor.writeheader()
        escritor.writerows(todos)
    print(f"Total: {len(todos)} registros em {args.saida}")


if __name__ == "__main__":
    main()
